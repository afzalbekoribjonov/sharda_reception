from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _safe(v: Any) -> str:
    if v is None:
        return ""
    return str(v)


def build_candidates_excel(
    rows: list[dict[str, Any]],
    include_credentials: bool,
    out_dir: Path,
) -> Path:
    """
    Creates an .xlsx file and returns its path.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    path = out_dir / f"candidates_export_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"

    headers = [
        "Full Name",
        "Phone",
        "Faculty",
        "B.Tech Track",
        "Exam Type",
        "Exam Date Time",
        "Address",
        "Location (lat,lng)",
        "Status",
        "Telegram ID",
    ]

    if include_credentials:
        headers.insert(7, "Online Link")
        headers.insert(8, "Login")
        headers.insert(9, "Password")

    ws.append(headers)

    for r in rows:
        loc = r.get("location") or {}
        loc_text = ""
        if isinstance(loc, dict) and loc.get("lat") is not None and loc.get("lng") is not None:
            loc_text = f"{loc.get('lat')},{loc.get('lng')}"

        base = [
            _safe(r.get("full_name")),
            _safe(r.get("phone")),
            _safe(r.get("faculty")),
            _safe(r.get("btech_track")),
            _safe(r.get("exam_type")),
            _safe(r.get("exam_date_time")),
            _safe(r.get("address")),
            loc_text,
            _safe(r.get("status")),
            _safe(r.get("telegram_id")),
        ]

        if include_credentials:
            # Insert credentials after address column logically
            # base currently: ... Address, Location, Status, TelegramID
            # We want: Address, OnlineLink, Login, Password, Location, Status, TelegramID
            address = base[6]
            location = base[7]
            status = base[8]
            tid = base[9]
            row = [
                base[0], base[1], base[2], base[3], base[4], base[5],
                address,
                _safe(r.get("online_link")),
                _safe(r.get("exam_login")),
                _safe(r.get("exam_password")),
                location,
                status,
                tid,
            ]
        else:
            row = base

        ws.append(row)

    # Auto column widths (simple)
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 45)

    wb.save(path)
    return path